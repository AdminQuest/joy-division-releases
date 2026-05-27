#!/usr/bin/env python3
"""Migre les XLSX legacy degonfles vers YAML variants + ownerships.

Pipeline en deux passes :
  1. Lit scripts/layouts.yml et scripts/audio_format_patterns.yml.
  2. Pre-scan : pour chaque XLSX, identifie les SERIES de sous-variantes
     (legacy_code partageant un prefixe, suffixes a/b/c/... uniques).
     Construit la table series_membership[(xlsx, row)] -> (group_id,
     group_role).
  3. Pass 2 : pour chaque ligne, produit le YAML variant (avec
     variant_group attache si applicable) + le YAML ownership
     (avec acquired_at omis, convention 2000-NNN) si owned == True.
     Valide chaque YAML immediatement contre son schema.

Decisions implementees (v2) :
  A) year omis quand absent dans le legacy (pas de fallback 2000).
  B) Reclassification BOOK -> PARA quand le titre est un magazine
     (#NN), fanzine ou catalogue d'enchere.
  C) Filtre des lignes d'en-tete BOOK (titre = "Titre" / "TITRE" /
     "Books" / "Livres" / "Auteur" / "Author").
  D) variant_group active pour les series BOOT a/b/c/d/e :
     - group_id derive du prefixe legacy_code (ex. "BW4")
     - group_role = color_variation si tous les membres partagent
       le canonical_title, sinon pressing_variation.
  E) Extraction de la couleur depuis la chaine format ET la colonne
     Version, avec priorite Version > format_str.
  F) Nouveaux patterns audio (Splatter LP, color LP, V1: N LP,
     emoji CD, Box Set -> coffret_skip).
  G) Split des matrices sur retours-ligne uniquement.
  H) Cleanup des sorties avant relance.

Conformement a la consigne, ce script ne commit aucun YAML : les
sorties restent locales pour relecture.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import yaml
from jsonschema import Draft202012Validator
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Constantes de chemins
# ---------------------------------------------------------------------------

REPO_RELEASES = Path("/home/user/joy-division-releases")
REPO_COLLECTION = Path("/home/user/joy-division-collection")
SLIM_DIR = Path("/home/user/_legacy_dump/slimmed")
REPORT_PATH = Path("/home/user/_legacy_dump/migration_report.md")
MAPPING_DIR = REPO_RELEASES / "data" / "_migration"
MAPPING_CSV = MAPPING_DIR / "legacy_mapping.csv"
VARIANTS_DIR = REPO_RELEASES / "data" / "variants"
OWNERSHIP_DIR = REPO_COLLECTION / "data" / "ownership"

LAYOUTS_YML = REPO_RELEASES / "scripts" / "layouts.yml"
PATTERNS_YML = REPO_RELEASES / "scripts" / "audio_format_patterns.yml"

VARIANT_SCHEMA_PATH = REPO_RELEASES / "schema" / "variant.schema.json"
OWNERSHIP_SCHEMA_PATH = REPO_COLLECTION / "schema" / "ownership.schema.json"


# ---------------------------------------------------------------------------
# Constantes metier
# ---------------------------------------------------------------------------

LANGUAGE_MAP = {
    "français": "fr", "francais": "fr", "fr": "fr",
    "anglais": "en", "english": "en", "en": "en",
    "allemand": "de", "deutsch": "de", "de": "de",
    "italien": "it", "italiano": "it", "it": "it",
    "espagnol": "es", "español": "es", "es": "es",
    "néerlandais": "nl", "nederlands": "nl", "nl": "nl",
    "portugais": "pt", "português": "pt", "pt": "pt",
    "polonais": "pl", "polski": "pl", "pl": "pl",
    "serbe": "sr", "srpski": "sr", "sr": "sr",
    "japonais": "ja", "ja": "ja",
}

ARTICLES = ["The ", "A ", "An ", "Le ", "La ", "Les ", "L'", "Un ", "Une "]

SUFFIX_PATTERN = re.compile(r"^(.+?)([a-j])$")

# Lignes d'en-tete a rejeter dans BOOK
BOOK_HEADER_TITLES = {
    "titre", "title", "books", "book", "livres", "livre",
    "auteur", "author", "authors", "editeur", "publisher",
    "langue", "language", "annee", "année", "year",
    "type", "isbn", "pays", "pages",
}

# Patterns pour reclassifier BOOK -> PARA
MAGAZINE_ISSUE_RE = re.compile(r"#\s*\d+")
FANZINE_RE = re.compile(r"(?i)\bfanzine\b")
MAGAZINE_KEYWORD_RE = re.compile(r"(?i)\bmagazine\b")
AUCTION_RE = re.compile(r"(?i)auction\s+catal")


# ---------------------------------------------------------------------------
# Categorisation par nom de fichier (decision 6 — locked)
# ---------------------------------------------------------------------------


def categorize(filename: str) -> tuple[str, str | None]:
    if filename.startswith("Bootlegs - "):
        return ("bootleg", "BOOT")
    if filename == "Officiels et pirates - VHS & DVD.xlsx":
        return ("video", "VID")
    if filename == "Officiels et pirates - Coffrets.xlsx":
        return ("coffret_skip", None)
    if filename.startswith("Officiels et pirates - "):
        return ("officiel_or_pirate", "OFF_or_PIR")
    if filename.startswith("Livres -"):
        return ("livre_or_para", "BOOK_or_PARA")
    return ("unknown", None)


# ---------------------------------------------------------------------------
# Lettre du variant_id
# ---------------------------------------------------------------------------


def compute_letter(title: str | None) -> str | None:
    if not title:
        return None
    cleaned = str(title).strip()
    if not cleaned:
        return None
    for art in ARTICLES:
        if cleaned.lower().startswith(art.lower()):
            cleaned = cleaned[len(art):].strip()
            break
    if not cleaned:
        return None
    first = cleaned[0].upper()
    if "A" <= first <= "Z":
        return first
    return "0"


# ---------------------------------------------------------------------------
# Cellules
# ---------------------------------------------------------------------------


def cell_value(ws, row: int, col: int):
    if col is None:
        return None
    return ws.cell(row=row, column=col).value


def cell_link(ws, row: int, col: int) -> str | None:
    if col is None:
        return None
    cell = ws.cell(row=row, column=col)
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def parse_year(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int) and 1976 <= value <= 2030:
        return value
    s = str(value)
    m = re.search(r"\b(19[7-9][0-9]|20[0-3][0-9])\b", s)
    if not m:
        return None
    y = int(m.group(1))
    if 1976 <= y <= 2030:
        return y
    return None


def parse_price_eur(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value > 0 else None
    s = str(value).strip()
    if not s or s.startswith("/") or s.startswith("#"):
        return None
    m = re.match(r"^\s*(\d+(?:[.,]\d+)?)", s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", "."))
    except ValueError:
        return None


def parse_copy_number(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or len(s) > 40:
        return None
    return s


def normalize_country(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"[^\x00-\x7F]", "", s).strip()
    return s or None


# ---------------------------------------------------------------------------
# Color extraction (decision E)
# ---------------------------------------------------------------------------


def build_color_regex(tokens: list[str]) -> re.Pattern:
    escaped = "|".join(re.escape(t) for t in sorted(tokens, key=len, reverse=True))
    # Match a contiguous sequence of color tokens.
    return re.compile(rf"\b((?:{escaped})(?:\s+(?:{escaped}))*)\b", re.IGNORECASE)


def extract_color_from_string(s: str | None, color_re: re.Pattern) -> str | None:
    if not s:
        return None
    text = str(s).strip()
    if not text:
        return None
    m = color_re.search(text)
    if not m:
        return None
    matched = m.group(1).strip()
    if not matched:
        return None
    return matched


# ---------------------------------------------------------------------------
# Audio / video parser
# ---------------------------------------------------------------------------


class AudioFormatParser:
    def __init__(self, patterns_data: dict):
        self.patterns = [
            {"regex": re.compile(e["regex"]), "result": e["result"]}
            for e in patterns_data.get("patterns", [])
        ]
        self.mixed_patterns = [
            re.compile(p) for p in patterns_data.get("mixed_audio_video_patterns", [])
        ]
        self.box_set_patterns = [
            re.compile(p) for p in patterns_data.get("box_set_patterns", [])
        ]
        self.video_patterns = [
            {"regex": re.compile(e["regex"]), "support_type": e["support_type"]}
            for e in patterns_data.get("video_patterns", [])
        ]
        self.color_re = build_color_regex(patterns_data.get("color_tokens", []))

    def is_mixed_audio_video(self, format_str: str | None) -> bool:
        if not format_str:
            return False
        return any(p.search(str(format_str)) for p in self.mixed_patterns)

    def is_box_set(self, format_str: str | None) -> bool:
        if not format_str:
            return False
        return any(p.search(str(format_str)) for p in self.box_set_patterns)

    def parse_audio(
        self, format_str: str | None, version_str: str | None
    ) -> tuple[dict, bool]:
        """Retourne (support_dict, matched_cleanly)."""
        if not format_str:
            return ({"support_type": "vinyl", "count": 1}, False)
        s = str(format_str).strip()
        matched = False
        result: dict = {"support_type": "vinyl", "count": 1}
        for entry in self.patterns:
            m = entry["regex"].search(s)
            if m:
                result = {}
                base = entry["result"]
                for k, v in base.items():
                    if k == "count_group":
                        try:
                            cg = m.group(v)
                            result["count"] = int(cg) if cg else 1
                        except (IndexError, ValueError, TypeError):
                            result["count"] = 1
                    elif k == "color_default":
                        # Place tentatif, ecrase par extraction explicite ci-apres.
                        result["color"] = v
                    else:
                        result[k] = v
                if "count" not in result:
                    result["count"] = 1
                matched = True
                break

        # Color extraction : Version > format_str
        color = extract_color_from_string(version_str, self.color_re)
        if not color:
            color = extract_color_from_string(format_str, self.color_re)
        if color:
            result["color"] = color
        elif "color" not in result and version_str:
            # Fallback : Version libre non matchee par color_re
            vs = str(version_str).strip()
            if vs and len(vs) <= 60:
                result["color"] = vs

        return (result, matched)

    def parse_video(self, format_str: str | None) -> tuple[dict, bool]:
        if not format_str:
            return ({"support_type": "dvd", "count": 1}, False)
        s = str(format_str)
        for entry in self.video_patterns:
            if entry["regex"].search(s):
                count = 1
                m = re.match(r"^\s*(\d+)\s*[xX]?", s)
                if m:
                    try:
                        count = int(m.group(1))
                    except ValueError:
                        count = 1
                return ({"support_type": entry["support_type"], "count": count}, True)
        return ({"support_type": "dvd", "count": 1}, False)


# ---------------------------------------------------------------------------
# Legacy code
# ---------------------------------------------------------------------------


def build_legacy_code(ws, row: int, layout: dict) -> dict:
    start = layout.get("legacy_code_start")
    if start is None:
        return {
            "segments": [],
            "reconstructed": "",
            "has_suffix": False,
            "source_row": row,
        }
    segments = []
    for c in range(start, start + 4):
        v = cell_value(ws, row, c)
        segments.append(None if v is None else str(v).strip())
    last = segments[-1] if segments else None
    has_suffix = bool(last and SUFFIX_PATTERN.match(last) and len(last) == 1)
    reconstructed = "".join(s for s in segments if s)
    return {
        "segments": segments,
        "reconstructed": reconstructed,
        "has_suffix": has_suffix,
        "source_row": row,
    }


# ---------------------------------------------------------------------------
# Determination O/P
# ---------------------------------------------------------------------------


def determine_op(ws, row: int, op_col: int | None) -> tuple[str, str, bool]:
    if op_col is None:
        return ("officiel", "OFF", False)
    val = cell_value(ws, row, op_col)
    if val is None:
        return ("officiel", "OFF", True)
    s = str(val).strip().upper()
    if s.startswith("O"):
        return ("officiel", "OFF", False)
    if s.startswith("P"):
        return ("pirate", "PIR", False)
    return ("officiel", "OFF", True)


# ---------------------------------------------------------------------------
# Reclassification BOOK -> PARA (decision B) + filtre header (decision C)
# ---------------------------------------------------------------------------


def is_book_header_row(title: str) -> bool:
    """Detecte une ligne d'en-tete ou de section dans Livres -.xlsx."""
    if not title:
        return True
    t = title.strip()
    if not t:
        return True
    # Si fait uniquement de separateurs / blancs
    if re.fullmatch(r"[\s\-_=*]+", t):
        return True
    if t.lower() in BOOK_HEADER_TITLES:
        return True
    # Sections style "Livres en français", "Livres en anglais"
    if re.match(r"(?i)^livres?\s+en\s+\w+$", t):
        return True
    return False


def reclassify_book(title: str) -> tuple[str, str, dict | None]:
    """Renvoie (release_type, prefix, para_extras_or_None).

    Si le livre est en realite un magazine / fanzine / catalogue,
    retourne ('para', 'PARA', {parent_object, parent_title, parent_issue}).
    Sinon ('livre', 'BOOK', None).
    """
    if not title:
        return ("livre", "BOOK", None)
    t = title.strip()

    # Auction catalogue -> misc
    if AUCTION_RE.search(t):
        return ("para", "PARA", {
            "parent_object": "misc",
            "parent_title": t,
            "object_description": f"Auction catalogue: {t}",
        })

    # Fanzine
    if FANZINE_RE.search(t):
        return ("para", "PARA", {
            "parent_object": "fanzine",
            "parent_title": t,
            "object_description": f"Fanzine: {t}",
        })

    # Magazine avec numero #NN
    m_issue = MAGAZINE_ISSUE_RE.search(t)
    if m_issue:
        issue = m_issue.group(0).strip()
        parent_title = MAGAZINE_ISSUE_RE.sub("", t).strip()
        return ("para", "PARA", {
            "parent_object": "magazine",
            "parent_title": parent_title or t,
            "parent_issue": issue,
            "object_description": f"Magazine: {t}",
        })

    # Mot-cle "magazine" dans le titre
    if MAGAZINE_KEYWORD_RE.search(t):
        return ("para", "PARA", {
            "parent_object": "magazine",
            "parent_title": t,
            "object_description": f"Magazine: {t}",
        })

    return ("livre", "BOOK", None)


# ---------------------------------------------------------------------------
# Builders format_details
# ---------------------------------------------------------------------------


def build_variant_audio(
    ws, row: int, layout: dict, parser: AudioFormatParser,
) -> tuple[dict, bool, list[str]]:
    issues: list[str] = []
    format_str = cell_value(ws, row, layout.get("format"))
    version_str = cell_value(ws, row, layout.get("version"))
    label_str = cell_value(ws, row, layout.get("label"))
    matrices_str = cell_value(ws, row, layout.get("matrices"))

    support, matched = parser.parse_audio(format_str, version_str)
    if not matched:
        issues.append(f"format audio non parse: {format_str!r}")

    audio_format: dict = {
        "type": "audio",
        "supports": [support],
    }
    if label_str:
        s = str(label_str).strip()
        if s and len(s) <= 200:
            audio_format["label_apparent"] = s

    # Decision G : split matrices sur retours-ligne uniquement
    if matrices_str:
        matrices = [
            m.strip()
            for m in re.split(r"[\r\n]+", str(matrices_str))
            if m.strip()
        ]
        if matrices:
            support["matrices"] = matrices[:20]

    return (audio_format, matched, issues)


def build_variant_video(
    ws, row: int, layout: dict, parser: AudioFormatParser,
) -> tuple[dict, bool, list[str]]:
    issues: list[str] = []
    format_str = cell_value(ws, row, layout.get("format"))
    support, matched = parser.parse_video(format_str)
    if not matched:
        issues.append(f"format video non parse: {format_str!r}")
    video_format = {"type": "video", "supports": [support]}
    return (video_format, matched, issues)


def build_variant_book(ws, row: int, layout: dict) -> tuple[dict, list[str]]:
    issues: list[str] = []
    author = cell_value(ws, row, layout.get("author"))
    publisher = cell_value(ws, row, layout.get("publisher"))
    language_raw = cell_value(ws, row, layout.get("language"))
    year_raw = cell_value(ws, row, layout.get("year"))
    pages_raw = cell_value(ws, row, layout.get("pages"))
    isbn = cell_value(ws, row, layout.get("isbn"))

    if author:
        authors = [
            a.strip()
            for a in re.split(r"\s*(?:,|&|\bet\b)\s*", str(author))
            if a.strip()
        ]
        if not authors:
            authors = ["Inconnu"]
            issues.append("auteur vide -> 'Inconnu'")
    else:
        authors = ["Inconnu"]
        issues.append("auteur non renseigne -> 'Inconnu'")

    if not publisher:
        publisher = "Inconnu"
        issues.append("editeur non renseigne -> 'Inconnu'")
    else:
        publisher = str(publisher).strip()

    pub_year = parse_year(year_raw)
    if pub_year is None:
        pub_year = 1900  # sentinel hors fenetre pour book_format (req >= 1900)
        issues.append("annee de publication non parsee -> 1900 (sentinel)")

    lang_norm = "fr"
    if language_raw:
        key = str(language_raw).strip().lower()
        lang_norm = LANGUAGE_MAP.get(key, "fr")
        if key not in LANGUAGE_MAP:
            issues.append(f"langue non mappee: {language_raw!r} -> 'fr'")

    book_format = {
        "type": "book",
        "authors": authors,
        "publisher": publisher,
        "publication_year": pub_year,
        "language": lang_norm,
        "format_physical": "softcover",
        "illustrations": True,
        "has_appendices": False,
    }

    if isbn:
        isbn_str = str(isbn).strip()
        m13 = re.match(r"^(97[89])[-\s]?([\d\-]+)$", isbn_str)
        if m13:
            book_format["isbn_13"] = f"{m13.group(1)}-{m13.group(2).replace(' ', '')}"
        elif re.match(r"^[\dX\-]+$", isbn_str):
            book_format["isbn_10"] = isbn_str.replace(" ", "")

    if pages_raw is not None:
        try:
            book_format["page_count"] = int(pages_raw)
        except (TypeError, ValueError):
            pass

    return (book_format, issues)


def build_variant_para(extras: dict) -> dict:
    """Construit para_format pour un magazine/fanzine reclassifie."""
    para = {
        "type": "para",
        "parent_object": extras["parent_object"],
        "object_description": extras["object_description"],
        "embedded_supports": [],
    }
    if "parent_title" in extras:
        para["parent_title"] = extras["parent_title"]
    if "parent_issue" in extras:
        para["parent_issue"] = extras["parent_issue"]
    if "parent_year" in extras:
        para["parent_year"] = extras["parent_year"]
    return para


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


class Validators:
    def __init__(self):
        variant_schema = json.loads(VARIANT_SCHEMA_PATH.read_text())
        ownership_schema = json.loads(OWNERSHIP_SCHEMA_PATH.read_text())
        Draft202012Validator.check_schema(variant_schema)
        Draft202012Validator.check_schema(ownership_schema)
        self.variant = Draft202012Validator(
            variant_schema, format_checker=Draft202012Validator.FORMAT_CHECKER
        )
        self.ownership = Draft202012Validator(
            ownership_schema, format_checker=Draft202012Validator.FORMAT_CHECKER
        )

    def check_variant(self, data: dict) -> list[str]:
        return [
            f"{list(e.absolute_path)}: {e.message}"
            for e in self.variant.iter_errors(data)
        ]

    def check_ownership(self, data: dict) -> list[str]:
        return [
            f"{list(e.absolute_path)}: {e.message}"
            for e in self.ownership.iter_errors(data)
        ]


# ---------------------------------------------------------------------------
# Detection lignes vides
# ---------------------------------------------------------------------------


def is_empty_row(ws, row: int, layout: dict) -> bool:
    checks = [layout.get("owned"), layout.get("title"), layout.get("legacy_code_start")]
    for c in checks:
        if c is not None and cell_value(ws, row, c) is not None:
            return False
    return True


# ---------------------------------------------------------------------------
# Pass 1 — detection des series de sous-variantes (decision D)
# ---------------------------------------------------------------------------


def detect_series(layouts: dict, xlsx_files: list[Path]) -> dict:
    """Renvoie series_membership[(xlsx_name, row)] -> (group_id, group_role).

    Une serie est un ensemble de >=2 lignes dans un meme XLSX dont les
    legacy_code partagent un prefixe commun et ont chacun un suffixe
    different parmi {a,b,c,...,j}.
    """
    series_membership: dict[tuple[str, int], tuple[str, str]] = {}

    for xlsx_path in xlsx_files:
        name = xlsx_path.name
        layout = layouts.get(name)
        if not layout:
            continue
        category, _ = categorize(name)
        if category != "bootleg":
            continue  # decision D : seul BOOT a des series a/b/c/d/e

        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active

        # Recolte les triplets (row, prefix, suffix, title)
        candidates = []
        empty_streak = 0
        for row in range(2, ws.max_row + 1):
            if is_empty_row(ws, row, layout):
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0
            code = build_legacy_code(ws, row, layout)
            reconstructed = code["reconstructed"]
            m = SUFFIX_PATTERN.match(reconstructed)
            if not m:
                continue
            prefix = m.group(1)
            suffix = m.group(2)
            title_val = cell_value(ws, row, layout.get("title"))
            title = str(title_val).strip() if title_val else ""
            candidates.append((row, prefix, suffix, title))
        wb.close()

        # Groupe par prefix
        groups: dict[str, list[tuple]] = defaultdict(list)
        for row, prefix, suffix, title in candidates:
            groups[prefix].append((row, suffix, title))

        for prefix, members in groups.items():
            if len(members) < 2:
                continue
            suffixes_set = {m[1] for m in members}
            if not all(s in "abcdefghij" for s in suffixes_set):
                continue
            titles = {m[2] for m in members if m[2]}
            if len(titles) <= 1:
                group_role = "color_variation"
            else:
                group_role = "pressing_variation"
            for row, _, _ in members:
                series_membership[(name, row)] = (prefix, group_role)

    return series_membership


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    # Cleanup (decision H)
    for d in (VARIANTS_DIR, OWNERSHIP_DIR, MAPPING_DIR):
        d.mkdir(parents=True, exist_ok=True)
    for f in VARIANTS_DIR.glob("*.yml"):
        f.unlink()
    for f in OWNERSHIP_DIR.glob("*.yml"):
        f.unlink()
    if MAPPING_CSV.exists():
        MAPPING_CSV.unlink()

    layouts_doc = yaml.safe_load(LAYOUTS_YML.read_text())
    patterns_doc = yaml.safe_load(PATTERNS_YML.read_text())
    parser = AudioFormatParser(patterns_doc)
    validators = Validators()

    layouts = layouts_doc.get("layouts", {})
    xlsx_files = sorted(SLIM_DIR.glob("*.xlsx"))

    # Pass 1 — series detection
    print("[Pass 1] detection des series BOOT...")
    series_membership = detect_series(layouts, xlsx_files)
    print(f"[Pass 1] {len({v[0] for v in series_membership.values()})} groupes detectes, {len(series_membership)} membres au total\n")

    # Compteurs
    variant_seq: dict[tuple[str, str], int] = defaultdict(int)
    ownership_seq: dict[int, int] = defaultdict(int)
    variants_by_prefix: dict[str, int] = defaultdict(int)
    ownerships_by_prefix: dict[str, int] = defaultdict(int)
    boot_letter_count: dict[str, int] = defaultdict(int)
    quality_count: dict[str, int] = defaultdict(int)
    files_count: dict[str, int] = defaultdict(int)

    rejects: list[dict] = []
    needs_review_list: list[dict] = []
    coffret_skipped: list[dict] = []
    reclassified_para: list[dict] = []
    unknown_layout_files: list[dict] = []
    mapping_rows: list[dict] = []

    total_variants = 0
    total_ownerships = 0
    total_xlsx = 0
    grouped_variants = 0

    for xlsx_path in xlsx_files:
        name = xlsx_path.name
        layout = layouts.get(name)
        if not layout:
            unknown_layout_files.append({"file": name, "reason": "absent de layouts.yml"})
            continue
        category, _ = categorize(name)
        if category == "unknown":
            unknown_layout_files.append({"file": name, "reason": "category=unknown"})
            continue

        print(f"[XLSX] {name}  (layout={layout.get('layout_id')}, cat={category})")
        total_xlsx += 1

        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        empty_streak = 0

        for row in range(2, max_row + 1):
            if is_empty_row(ws, row, layout):
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0

            # --- coffret_skip via Coffrets.xlsx ---
            if category == "coffret_skip":
                title_val = cell_value(ws, row, layout.get("title"))
                title_str = str(title_val).strip() if title_val else None
                coffret_skipped.append({
                    "file": name, "row": row,
                    "title": title_str or "(titre vide)",
                    "reason": "fichier Coffrets.xlsx",
                })
                continue

            # --- titre ---
            title_val = cell_value(ws, row, layout.get("title"))
            title = str(title_val).strip() if title_val else ""

            # --- filtre ligne d'en-tete BOOK (decision C) ---
            if category == "livre_or_para" and is_book_header_row(title):
                rejects.append({
                    "file": name, "row": row,
                    "reason": "ligne d'en-tete / section",
                })
                continue

            if not title:
                rejects.append({
                    "file": name, "row": row,
                    "reason": "titre vide",
                })
                continue

            letter = compute_letter(title)
            if letter is None:
                rejects.append({
                    "file": name, "row": row,
                    "reason": "lettre indeterminable",
                })
                continue

            # --- determination release_type et prefix ---
            issues: list[str] = []
            para_extras: dict | None = None
            if category == "bootleg":
                release_type, prefix = "bootleg", "BOOT"
            elif category == "video":
                release_type, prefix = "video", "VID"
            elif category == "livre_or_para":
                # Decision B : reclassification potentielle
                release_type, prefix, para_extras = reclassify_book(title)
                if para_extras:
                    reclassified_para.append({
                        "file": name, "row": row,
                        "title": title,
                        "parent_object": para_extras["parent_object"],
                    })
            elif category == "officiel_or_pirate":
                release_type, prefix, op_uncertain = determine_op(
                    ws, row, layout.get("op_column")
                )
                if op_uncertain:
                    issues.append("cellule O/P vide ou non reconnue -> OFF par defaut")
            else:
                rejects.append({"file": name, "row": row, "reason": f"category={category}"})
                continue

            # --- formats mixtes / box set -> coffret_skip ---
            if category in ("bootleg", "officiel_or_pirate"):
                fmt_str = cell_value(ws, row, layout.get("format"))
                if fmt_str:
                    if parser.is_mixed_audio_video(str(fmt_str)):
                        coffret_skipped.append({
                            "file": name, "row": row,
                            "title": title,
                            "reason": "format mixte audio+video",
                        })
                        continue
                    if parser.is_box_set(str(fmt_str)):
                        coffret_skipped.append({
                            "file": name, "row": row,
                            "title": title,
                            "reason": "format box set",
                        })
                        continue

            # --- numerotation ---
            variant_seq[(prefix, letter)] += 1
            seq = variant_seq[(prefix, letter)]
            variant_id = f"{prefix}-{letter}-{seq:04d}"

            # --- annee (decision A : pas de fallback) ---
            year = parse_year(cell_value(ws, row, layout.get("year")))

            # --- legacy_code ---
            legacy_code = build_legacy_code(ws, row, layout)
            legacy_code["source_xlsx"] = name

            # --- format_details ---
            cleanly_parsed = True
            if release_type in ("bootleg", "officiel", "pirate"):
                fd, matched, fmt_issues = build_variant_audio(ws, row, layout, parser)
                if not matched:
                    cleanly_parsed = False
                    issues.extend(fmt_issues)
            elif release_type == "video":
                fd, matched, fmt_issues = build_variant_video(ws, row, layout, parser)
                if not matched:
                    cleanly_parsed = False
                    issues.extend(fmt_issues)
            elif release_type == "livre":
                fd, book_issues = build_variant_book(ws, row, layout)
                issues.extend(book_issues)
                cleanly_parsed = False  # livres toujours en needs_review
            elif release_type == "para":
                # PARA reclassifie : on essaie de recuperer parent_year
                if year:
                    para_extras["parent_year"] = year
                fd = build_variant_para(para_extras)
                cleanly_parsed = False  # para reclassifie toujours en needs_review
                issues.append(f"reclassifie BOOK->PARA ({para_extras['parent_object']})")
            else:
                rejects.append({"file": name, "row": row, "reason": f"release_type={release_type}"})
                variant_seq[(prefix, letter)] -= 1
                continue

            # --- documentation_quality ---
            if release_type in ("livre", "para") or not cleanly_parsed or issues:
                doc_quality = "needs_review"
            else:
                doc_quality = "verified"

            # --- liens externes ---
            joydiv_url = cell_link(ws, row, layout.get("title"))
            if not joydiv_url and layout.get("kind") == "audio_or_video":
                joydiv_url = cell_link(ws, row, layout.get("personal_1"))
            discogs_url = cell_link(ws, row, layout.get("discogs"))

            # --- country ---
            country = normalize_country(cell_value(ws, row, layout.get("country")))

            # --- notes ---
            notes = None
            if layout.get("kind") == "audio_or_video":
                comments = cell_value(ws, row, layout.get("comments"))
                if comments:
                    notes = str(comments).strip()[:1000] or None
            elif layout.get("kind") == "book":
                summary = cell_value(ws, row, layout.get("summary"))
                if summary:
                    notes = str(summary).strip()[:1000] or None

            # --- variant_group (decision D) ---
            variant_group = None
            if prefix == "BOOT" and (name, row) in series_membership:
                gid, role = series_membership[(name, row)]
                variant_group = {
                    "group_id": gid,
                    "group_role": role,
                }
                grouped_variants += 1

            # --- variant doc ---
            variant_doc: dict = {
                "variant_id": variant_id,
                "release_type": release_type,
                "canonical_title": title,
                "canonical_artist": "Joy Division",
                "documentation_quality": doc_quality,
                "format_details": fd,
                "joydiv_letter": letter,
                "_legacy_code": legacy_code,
            }
            if year is not None:
                variant_doc["year"] = year
            if joydiv_url:
                variant_doc["joydiv_url"] = joydiv_url
            if discogs_url:
                variant_doc["discogs_url"] = discogs_url
            if country:
                variant_doc["country_or_pressing_place"] = country
            if notes:
                variant_doc["notes"] = notes
            if variant_group:
                variant_doc["variant_group"] = variant_group

            # Ordonne les cles : tronc commun avant format_details/legacy
            ordered_keys = [
                "variant_id", "release_type", "canonical_title", "canonical_artist",
                "year", "documentation_quality", "joydiv_letter", "joydiv_url",
                "discogs_url", "country_or_pressing_place", "variant_group",
                "notes", "format_details", "_legacy_code",
            ]
            variant_doc = {k: variant_doc[k] for k in ordered_keys if k in variant_doc}

            # Validation
            v_errors = validators.check_variant(variant_doc)
            if v_errors:
                msg = (
                    f"\n[FATAL] validation variant {variant_id} ({name} row {row})"
                    f"\n  Title: {title!r}"
                    f"\n  Errors:"
                )
                for e in v_errors:
                    msg += f"\n    - {e}"
                print(msg, file=sys.stderr)
                return 1

            (VARIANTS_DIR / f"{variant_id}.yml").write_text(
                yaml.safe_dump(variant_doc, allow_unicode=True, sort_keys=False, width=120)
            )

            total_variants += 1
            variants_by_prefix[prefix] += 1
            quality_count[doc_quality] += 1
            files_count[name] += 1
            if prefix == "BOOT":
                boot_letter_count[letter] += 1
            if doc_quality == "needs_review":
                needs_review_list.append({
                    "variant_id": variant_id,
                    "source": f"{name} row {row}",
                    "reason": "; ".join(issues) if issues else "par defaut",
                })

            # --- ownership ---
            ownership_id = None
            owned_val = cell_value(ws, row, layout.get("owned"))
            owned = (str(owned_val).strip().upper() == "TRUE") if owned_val is not None else False
            if owned:
                year_acq = 2000
                ownership_seq[year_acq] += 1
                seq_o = ownership_seq[year_acq]
                if seq_o > 999:
                    year_acq += 1
                    ownership_seq[year_acq] += 1
                    seq_o = ownership_seq[year_acq]
                ownership_id = f"own-{year_acq:04d}-{seq_o:03d}"

                photo_link = None
                if layout.get("kind") == "audio_or_video":
                    photo_link = cell_link(ws, row, layout.get("personal_1"))
                    if not photo_link:
                        photo_link = cell_link(ws, row, layout.get("personal_2"))

                price_paid = parse_price_eur(cell_value(ws, row, layout.get("cote")))
                copy_number = parse_copy_number(cell_value(ws, row, layout.get("limited")))

                ownership_doc: dict = {
                    "ownership_id": ownership_id,
                    "variant_id": variant_id,
                }
                if price_paid is not None:
                    ownership_doc["price_paid_eur"] = price_paid
                if copy_number is not None:
                    ownership_doc["copy_number"] = copy_number
                if photo_link:
                    ownership_doc["personal_photos"] = [photo_link]
                if notes:
                    ownership_doc["notes"] = notes[:500]

                o_errors = validators.check_ownership(ownership_doc)
                if o_errors:
                    msg = (
                        f"\n[FATAL] validation ownership {ownership_id} ({name} row {row})"
                        f"\n  Errors:"
                    )
                    for e in o_errors:
                        msg += f"\n    - {e}"
                    print(msg, file=sys.stderr)
                    return 1

                (OWNERSHIP_DIR / f"{ownership_id}.yml").write_text(
                    yaml.safe_dump(ownership_doc, allow_unicode=True, sort_keys=False, width=120)
                )
                total_ownerships += 1
                ownerships_by_prefix[prefix] += 1

            mapping_rows.append({
                "source_xlsx": name,
                "source_row": row,
                "legacy_code_reconstructed": legacy_code["reconstructed"],
                "variant_id": variant_id,
                "ownership_id": ownership_id or "",
                "release_type": release_type,
                "title": title,
            })

        wb.close()

    # CSV
    with MAPPING_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "source_xlsx", "source_row", "legacy_code_reconstructed",
            "variant_id", "ownership_id", "release_type", "title",
        ])
        writer.writeheader()
        writer.writerows(mapping_rows)

    # Rapport
    write_report(
        total_xlsx=total_xlsx,
        total_variants=total_variants,
        total_ownerships=total_ownerships,
        grouped_variants=grouped_variants,
        series_count=len({v[0] for v in series_membership.values()}),
        coffret_skipped=coffret_skipped,
        rejects=rejects,
        reclassified_para=reclassified_para,
        variants_by_prefix=variants_by_prefix,
        ownerships_by_prefix=ownerships_by_prefix,
        boot_letter_count=boot_letter_count,
        quality_count=quality_count,
        files_count=files_count,
        needs_review_list=needs_review_list,
        unknown_layout_files=unknown_layout_files,
    )

    print(f"\n--- Migration terminee ---")
    print(f"Variants : {total_variants}")
    print(f"Ownerships : {total_ownerships}")
    print(f"Variants groupes (variant_group) : {grouped_variants}")
    print(f"Reclassifications BOOK->PARA : {len(reclassified_para)}")
    print(f"Coffrets skippes : {len(coffret_skipped)}")
    print(f"Rejets : {len(rejects)}")
    print(f"needs_review : {quality_count['needs_review']}")
    return 0


def write_report(**kw):
    now = dt.datetime.now().isoformat(timespec="seconds")
    out: list[str] = []
    out.append(f"# Rapport de migration v2 — {now}\n")
    out.append("## Synthese\n")
    out.append(f"- XLSX lus : **{kw['total_xlsx']}**")
    out.append(f"- Variantes produites : **{kw['total_variants']}**")
    out.append(f"- Ownerships produits : **{kw['total_ownerships']}**")
    out.append(f"- Series detectees (variant_group) : **{kw['series_count']}**")
    out.append(f"- Variants membres d'une serie : **{kw['grouped_variants']}**")
    out.append(f"- Reclassifications BOOK -> PARA : **{len(kw['reclassified_para'])}**")
    out.append(f"- Coffrets identifies (skippes) : **{len(kw['coffret_skipped'])}**")
    out.append(f"- Lignes XLSX rejetees : **{len(kw['rejects'])}**")
    out.append(f"- Erreurs de validation : **0** (sinon le script aurait stoppe)")
    out.append("")

    out.append("## Decomposition par prefixe\n")
    out.append("| Prefixe | Variantes | Ownerships |")
    out.append("|---|---:|---:|")
    for pfx in ["BOOT", "OFF", "PIR", "VID", "BOX", "BOOK", "PARA"]:
        out.append(f"| {pfx} | {kw['variants_by_prefix'].get(pfx, 0)} | {kw['ownerships_by_prefix'].get(pfx, 0)} |")
    out.append("")

    out.append("## Decomposition BOOT par lettre\n")
    out.append("| Lettre | Variantes |")
    out.append("|---|---:|")
    for letter in ["0"] + [chr(c) for c in range(ord("A"), ord("Z") + 1)]:
        out.append(f"| {letter} | {kw['boot_letter_count'].get(letter, 0)} |")
    out.append("")

    out.append("## Qualite de documentation\n")
    out.append("| Statut | Nombre |")
    out.append("|---|---:|")
    for q in ["verified", "needs_review", "stub"]:
        out.append(f"| {q} | {kw['quality_count'].get(q, 0)} |")
    out.append("")

    out.append("## Top 10 XLSX par variantes produites\n")
    out.append("| XLSX | Variantes |")
    out.append("|---|---:|")
    for fname, n in sorted(kw['files_count'].items(), key=lambda x: -x[1])[:10]:
        out.append(f"| {fname} | {n} |")
    out.append("")

    out.append(f"## Reclassifications BOOK -> PARA ({len(kw['reclassified_para'])})\n")
    if kw['reclassified_para']:
        out.append("| XLSX | Ligne | Titre | parent_object |")
        out.append("|---|---:|---|---|")
        for r in kw['reclassified_para'][:200]:
            t = r['title'].replace('|', '\\|')[:80]
            out.append(f"| {r['file']} | {r['row']} | {t} | {r['parent_object']} |")
    else:
        out.append("Aucune.")
    out.append("")

    out.append(f"## Lignes rejetees ({len(kw['rejects'])})\n")
    if kw['rejects']:
        out.append("| XLSX | Ligne | Raison |")
        out.append("|---|---:|---|")
        for r in kw['rejects'][:100]:
            out.append(f"| {r['file']} | {r['row']} | {r['reason']} |")
        if len(kw['rejects']) > 100:
            out.append(f"| ... | ... | (+{len(kw['rejects'])-100} autres) |")
    else:
        out.append("Aucune.")
    out.append("")

    out.append(f"## Coffrets identifies pour seconde passe ({len(kw['coffret_skipped'])})\n")
    if kw['coffret_skipped']:
        out.append("| XLSX | Ligne | Titre | Raison |")
        out.append("|---|---:|---|---|")
        for c in kw['coffret_skipped'][:100]:
            t = c.get('title', '?').replace('|', '\\|')[:60]
            out.append(f"| {c['file']} | {c['row']} | {t} | {c.get('reason', '?')} |")
        if len(kw['coffret_skipped']) > 100:
            out.append(f"| ... | ... | ... | (+{len(kw['coffret_skipped'])-100} autres) |")
    else:
        out.append("Aucun.")
    out.append("")

    out.append(f"## Variants en needs_review ({len(kw['needs_review_list'])})\n")
    if kw['needs_review_list']:
        out.append("| variant_id | source | raison |")
        out.append("|---|---|---|")
        for r in kw['needs_review_list'][:150]:
            reason = r['reason'].replace('|', '\\|')[:200]
            out.append(f"| `{r['variant_id']}` | {r['source']} | {reason} |")
        if len(kw['needs_review_list']) > 150:
            out.append(f"| ... | ... | (+{len(kw['needs_review_list'])-150} autres) |")
    else:
        out.append("Aucun.")
    out.append("")

    out.append("## XLSX non traites (layout inconnu)\n")
    if kw['unknown_layout_files']:
        out.append("| XLSX | Raison |")
        out.append("|---|---|")
        for u in kw['unknown_layout_files']:
            out.append(f"| {u['file']} | {u['reason']} |")
    else:
        out.append("Aucun.")
    out.append("")

    REPORT_PATH.write_text("\n".join(out) + "\n")
    print(f"Rapport ecrit : {REPORT_PATH}")


if __name__ == "__main__":
    sys.exit(main())
